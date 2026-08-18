"""Extract SAQA NLRD "All Qualifications and Part-Qualifications" xlsx rows into
validated SaqaQualification models. Independent of the DHET register pipeline —
no institution-matching happens here (see web/lib/qualificationsMatching.ts)."""

import re

import openpyxl

from models import SaqaQualification

_NQF_LEVEL_RE = re.compile(r"NQF Level\s*(\d+)", re.IGNORECASE)
_TRAILING_LEVEL_RE = re.compile(r"L(\d+)\s*$")


def parse_nqf_level(raw):
    """Parses the SAQA "NQF Level" column, which is a free-text string rather
    than a number, e.g. "NQF Level 06" -> 6, "Level TBA: Pre-2009 was L4" -> 4,
    "Not Applicable" -> None."""
    if not raw:
        return None
    match = _NQF_LEVEL_RE.search(raw)
    if match:
        return int(match.group(1))
    match = _TRAILING_LEVEL_RE.search(raw.strip())
    if match:
        return int(match.group(1))
    return None


def iter_qualification_rows(xlsx_path):
    """Yields header-mapped row dicts for every qualification row in the SAQA NLRD
    register, across all NQF sub-frameworks (HEQSF, OQSF, GFETQSF, SFAP, SFNA) — the
    ingestion layer captures everything; filtering to a specific sub-framework (e.g.
    HEQSF-only for a higher-education product) is a consumer-time concern, not
    something dropped at read time."""
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)

    header = next(rows)
    columns = [str(cell).strip() if cell is not None else "" for cell in header]

    for values in rows:
        if not values:
            continue
        yield dict(zip(columns, values))


def row_to_qualification(row):
    """Returns a SaqaQualification, or None if the row is missing a qual id,
    title, or originator (mirrors build.record_to_institution's skip-don't-raise
    convention for malformed rows)."""
    qual_id = row.get("Qual ID")
    title = (row.get("Qualification Title") or "").strip()
    originator = (row.get("Originator") or "").strip()
    subfield = (row.get("Subfield") or "").strip()
    framework = (row.get("NQF Sub-Framework") or "").strip()

    if not qual_id or not title or not originator:
        return None

    nqf_level_raw = row.get("NQF Level") or ""
    credits = row.get("Min Credits")

    return SaqaQualification(
        qualId=int(qual_id),
        title=title,
        nqfLevel=parse_nqf_level(nqf_level_raw),
        nqfLevelRaw=nqf_level_raw,
        credits=int(credits) if credits else None,
        subfield=subfield,
        originator=originator,
        framework=framework,
    )


def build_qualifications(xlsx_path):
    """The single entry point: reads the SAQA xlsx and returns a flat list of
    SaqaQualification models across all NQF sub-frameworks (mirrors
    build.build_institutions). Callers that need a single sub-framework (e.g.
    HEQSF-only) filter on `.framework` themselves."""
    qualifications = []
    for row in iter_qualification_rows(xlsx_path):
        qualification = row_to_qualification(row)
        if qualification is not None:
            qualifications.append(qualification)
    return qualifications
